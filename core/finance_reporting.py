"""Builder dimension — Finance/Controlling management reporting (Dimensions / Phase 6).

Consolidate many internal documents full of figures into ONE management/board report where **every
number is traced to its source** (``playbooks/finance_reporting_playbook.md`` — zero hallucination).
Reads are local (Block A readers); the report is built by the local LLM and emitted as a Markdown
blueprint plus an Excel KPI table. Rendering to a Neura PDF/PPTX reuses Porter's ``prepare``.

All LLM access goes through :class:`LocalLLMClient` (RULE 6). Parsing is tolerant
(:func:`core.json_utils.extract_json_object`) and **fails open**: an unparsable response yields an
empty report with an explicit gap, never a crash.
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

from core.config import AppConfig
from core.json_utils import extract_json_object
from llm.local_llm_client import LocalLLMClient
from models.reporting import KeyFigure, ManagementReport, ReportSection
from models.research import DocContent

_PLAYBOOK_FILE = (
    Path(__file__).resolve().parent.parent / "playbooks" / "finance_reporting_playbook.md"
)
_MAX_TOTAL_CHARS = 16000  # total document budget fed to one consolidation call

_FALLBACK_RUBRIC = (
    "Consolidate the documents into one management report. Every figure MUST come from the "
    "documents, quoted exactly with its unit, period, and source. Never invent or round. Mark "
    "anything the documents do not answer as a gap."
)


def _slug(text: str) -> str:
    """Make a short, filesystem-safe slug from a title."""
    cleaned = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return (cleaned or "report")[:40]


def _argb(hex_color: str) -> str:
    """Normalize a ``#rrggbb`` color to openpyxl 8-digit ARGB."""
    value = hex_color.lstrip("#").upper()
    return value if len(value) == 8 else "FF" + value


def _load_rubric() -> str:
    """Load the finance playbook text (best-effort; fallback rubric if missing)."""
    try:
        text = _PLAYBOOK_FILE.read_text(encoding="utf-8").strip()
    except OSError:
        return _FALLBACK_RUBRIC
    return text or _FALLBACK_RUBRIC


def _documents_block(documents: list[DocContent], *, max_total: int = _MAX_TOTAL_CHARS) -> str:
    """Concatenate document texts with explicit [Source: name] markers, within a char budget."""
    parts: list[str] = []
    budget = max_total
    for document in documents:
        if budget <= 0:
            break
        name = Path(str(document.source_path)).name
        chunk = document.text[:budget]
        budget -= len(chunk)
        parts.append(f'[Source: {name}]\n"""{chunk}"""')
    return "\n\n".join(parts)


def build_report(
    llm: LocalLLMClient,
    documents: list[DocContent],
    *,
    period: str = "",
    title: str = "",
    language: str = "en",
) -> ManagementReport:
    """Consolidate documents into a :class:`ManagementReport` (every figure traced; fail-open)."""
    sources = [Path(str(d.source_path)).name for d in documents]
    rubric = _load_rubric()
    prompt = (
        f"PERIOD: {period or '(infer from the documents)'}\n"
        f"DOCUMENTS:\n{_documents_block(documents)}\n\n"
        "Return ONLY JSON:\n"
        '{"title": "<report title>", "period": "<reporting period>", '
        '"bottom_line": "<one-paragraph so-what>", '
        '"key_figures": [{"metric": "<name>", "value": "<exact value with unit>", '
        '"period": "<period>", "source": "<file/sheet/page>"}], '
        '"sections": [{"heading": "<so-what heading>", "bullets": ["..."]}], '
        '"risks": ["<risk · source>"], "gaps": ["<what the documents do NOT answer>"]}'
    )
    system = (
        f"{rubric}\n\nYou consolidate internal figures into one management report. "
        "Output only a single JSON object."
    )
    raw = llm.generate(prompt, system=system)
    data = extract_json_object(raw)
    if not data:
        return ManagementReport(
            title=title or "Management Report",
            period=period,
            sources=sources,
            language=language,
            gaps=["The model's report could not be parsed — re-run or review the documents."],
        )

    key_figures: list[KeyFigure] = []
    for item in data.get("key_figures", []) or []:
        if not isinstance(item, dict):
            continue
        metric = str(item.get("metric", "")).strip()
        if not metric:
            continue
        key_figures.append(
            KeyFigure(
                metric=metric,
                value=str(item.get("value", "")).strip(),
                period=str(item.get("period", "")).strip(),
                source=str(item.get("source", "")).strip(),
            )
        )

    sections: list[ReportSection] = []
    for item in data.get("sections", []) or []:
        if not isinstance(item, dict):
            continue
        heading = str(item.get("heading", "")).strip()
        if not heading:
            continue
        bullets = [str(b).strip() for b in (item.get("bullets") or []) if str(b).strip()]
        sections.append(ReportSection(heading=heading, bullets=bullets))

    return ManagementReport(
        title=title or str(data.get("title", "")).strip() or "Management Report",
        period=period or str(data.get("period", "")).strip(),
        bottom_line=str(data.get("bottom_line", "")).strip(),
        key_figures=key_figures,
        sections=sections,
        risks=[str(r).strip() for r in (data.get("risks") or []) if str(r).strip()],
        gaps=[str(g).strip() for g in (data.get("gaps") or []) if str(g).strip()],
        sources=sources,
        language=language,
    )


def render_report_markdown(report: ManagementReport) -> str:
    """Render the management report as a Markdown blueprint (the 'Spickzettel')."""
    lines = [f"# {report.title}" + (f" ({report.period})" if report.period else ""), ""]
    if report.bottom_line:
        lines += ["## Bottom Line", report.bottom_line, ""]
    if report.key_figures:
        lines += ["## Key Figures", "", "| Metric | Value | Period | Source |", "|---|---|---|---|"]
        for figure in report.key_figures:
            lines.append(
                f"| {figure.metric} | {figure.value} | {figure.period} | {figure.source} |"
            )
        lines.append("")
    for section in report.sections:
        lines.append(f"## {section.heading}")
        lines += [f"- {bullet}" for bullet in section.bullets]
        lines.append("")
    if report.risks:
        lines += ["## Risks & Dependencies"] + [f"- {risk}" for risk in report.risks] + [""]
    if report.gaps:
        lines += ["## Gaps & Data Quality"] + [f"- {gap}" for gap in report.gaps] + [""]
    if report.sources:
        lines += ["## Source Documents", ", ".join(report.sources), ""]
    return "\n".join(lines).strip()


def render_report_excel(report: ManagementReport, out_dir: Path | str, config: AppConfig) -> Path:
    """Write the key figures (with source) to a Neura-styled Excel workbook; return its path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    colors = config.output.colors
    header_fill = PatternFill("solid", fgColor=_argb(colors.excel_header))
    header_font = Font(bold=True, color=_argb(colors.white))
    wrap = Alignment(vertical="top", wrap_text=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Key Figures"
    worksheet["A1"] = report.title + (f" — {report.period}" if report.period else "")
    worksheet["A1"].font = Font(bold=True, size=14, color=_argb(colors.text_dark))
    worksheet["A2"] = (
        f"Generated {date.today().isoformat()} · Local · Every figure traced to its source."
    )
    worksheet["A2"].font = Font(italic=True, size=9, color=_argb(colors.charcoal))

    header_row = 4
    for column, title in enumerate(["Metric", "Value", "Period", "Source"], start=1):
        cell = worksheet.cell(row=header_row, column=column, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    row = header_row
    for figure in report.key_figures:
        row += 1
        for column, value in enumerate(
            [figure.metric, figure.value, figure.period, figure.source], start=1
        ):
            worksheet.cell(row=row, column=column, value=value).alignment = wrap

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 28

    output_dir = Path(out_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{date.today().isoformat()}_{_slug(report.title)}_key_figures.xlsx"
    path = output_dir / filename
    workbook.save(str(path))
    return path
