"""Excel output builder (SPEC §7 ``excel_builder.py``, §12): the four .xlsx templates via openpyxl.

Creates professional, **formula-driven** workbooks (N-10): every derived value is a real Excel
formula (``SUMPRODUCT``/``RANK``/``NPV``/…), never a hardcoded intermediate, so changing a yellow
input cell recalculates everything downstream when the file is opened in Microsoft Excel. Color
coding (yellow=input, blue=formula, green=positive, red=risk, dark=header) comes from
``config.output.colors`` (RULE 4). Fully local — openpyxl writes .xlsx with zero network (N-4).

Templates (SPEC §12): E-1 Decision/Scoring Matrix, E-2 Intelligence/Benchmark Table, E-3 Business
Case model, E-4 Tracker/Status Dashboard. Structured content is shaped by ``core.content_shaper``.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from core.config import AppConfig
from models.task import Language
from models.workbook import (
    BenchmarkData,
    BusinessCaseData,
    DecisionMatrixData,
    ExcelTemplate,
    TrackerData,
)


class ExcelBuildError(Exception):
    """An Excel workbook could not be built (fail fast with context)."""


def _slug(text: str) -> str:
    """Make a short, filesystem-safe slug from a title (shared with exporter naming)."""
    import re

    cleaned = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return (cleaned or "workbook")[:50]


def _t(language: Language, de: str, en: str) -> str:
    """Pick the German or English string for the language."""
    return de if language == Language.DE else en


def _argb(hex_color: str) -> str:
    """Normalize a ``#rrggbb`` / ``rrggbb`` color to openpyxl 8-digit ARGB."""
    value = hex_color.lstrip("#").upper()
    return value if len(value) == 8 else "FF" + value


class _Style:
    """Reusable openpyxl fills/fonts derived from the configured Neura palette (config-driven)."""

    def __init__(self, config: AppConfig) -> None:
        """Build the fills/fonts once from ``config.output.colors``."""
        c = config.output.colors
        self.input_fill = PatternFill("solid", fgColor=_argb(c.excel_input_cell))
        self.formula_fill = PatternFill("solid", fgColor=_argb(c.excel_formula_cell))
        self.positive_fill = PatternFill("solid", fgColor=_argb(c.excel_positive))
        self.negative_fill = PatternFill("solid", fgColor=_argb(c.excel_negative))
        self.header_fill = PatternFill("solid", fgColor=_argb(c.excel_header))
        self.surface_fill = PatternFill("solid", fgColor=_argb(c.light_surface))
        self.header_font = Font(name="Arial", bold=True, color=_argb(c.white), size=11)
        self.title_font = Font(name="Arial", bold=True, size=14, color=_argb(c.text_dark))
        self.bold = Font(name="Arial", bold=True, color=_argb(c.text_dark))
        self.normal = Font(name="Arial", color=_argb(c.text_dark))
        self.muted = Font(name="Arial", italic=True, size=9, color=_argb(c.charcoal))
        # 3-colour scale (low→high) for weighted scores, all from config.
        self.scale_low = _argb(c.excel_negative)
        self.scale_mid = _argb(c.excel_input_cell)
        self.scale_high = _argb(c.excel_positive)


_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _title_block(ws: Worksheet, style: _Style, title: str, instructions: str) -> None:
    """Write the universal A1 title / A2 last-updated / A3 instructions block (output_playbook)."""
    ws["A1"] = title
    ws["A1"].font = style.title_font
    ws["A2"] = f"{date.today().isoformat()}"
    ws["A2"].font = style.muted
    ws["A3"] = instructions
    ws["A3"].font = style.muted


# --------------------------------------------------------------- E-1 Decision / Scoring Matrix
_HEADER_ROW = 5
_WEIGHTS_ROW = 6
_FIRST_ENTITY_ROW = 7


def _normalized_weights(data: DecisionMatrixData) -> list[float]:
    """Return criterion weights normalized to sum to 1 (equal weights if none/zero)."""
    weights = [max(0.0, c.weight) for c in data.criteria]
    total = sum(weights)
    n = len(weights)
    if total <= 0 and n:
        return [1.0 / n] * n
    return [w / total for w in weights] if total else []


def _matrix_sheet(ws: Worksheet, style: _Style, data: DecisionMatrixData) -> None:
    """Write the Summary_Matrix tab: weights row (yellow) + SUMPRODUCT scores + RANK."""
    language = data.language
    n = len(data.criteria)
    last_crit_col = get_column_letter(1 + n)
    ws_col = 2 + n
    rank_col = 3 + n
    ws_letter = get_column_letter(ws_col)
    rank_letter = get_column_letter(rank_col)
    last_entity_row = _FIRST_ENTITY_ROW + len(data.entities) - 1

    _title_block(
        ws,
        style,
        data.title,
        _t(
            language,
            f"Gelbe Zellen sind Eingaben: Gewichte (Zeile {_WEIGHTS_ROW}) und Scores 1–5 "
            "anpassen — Gewichteter Score und Rang rechnen automatisch neu.",
            f"Yellow cells are inputs: change the weights (row {_WEIGHTS_ROW}) and the 1–5 "
            "scores — the Weighted Score and Rank recalculate automatically.",
        ),
    )

    # Header row.
    headers = [_t(language, "Unternehmen / Option", "Company / Option")]
    headers += [c.name for c in data.criteria]
    headers += [
        _t(language, "Gewichteter Score", "Weighted Score"),
        _t(language, "Rang", "Rank"),
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER

    # Weights row (yellow inputs) + auto-summing weight total.
    ws.cell(
        row=_WEIGHTS_ROW, column=1, value=_t(language, "Gewichtung", "Weight")
    ).font = style.bold
    for idx, weight in enumerate(_normalized_weights(data)):
        cell = ws.cell(row=_WEIGHTS_ROW, column=2 + idx, value=round(weight, 4))
        cell.fill = style.input_fill
        cell.number_format = "0%"
        cell.alignment = _CENTER
    total_cell = ws.cell(
        row=_WEIGHTS_ROW,
        column=ws_col,
        value=f"=SUM(B{_WEIGHTS_ROW}:{last_crit_col}{_WEIGHTS_ROW})",
    )
    total_cell.number_format = "0%"
    total_cell.font = style.bold
    total_cell.alignment = _CENTER

    # Entity rows: name + 1–5 scores + SUMPRODUCT weighted score + RANK.
    for offset, entity in enumerate(data.entities):
        row = _FIRST_ENTITY_ROW + offset
        ws.cell(row=row, column=1, value=entity.name).font = style.bold
        for idx in range(n):
            score = entity.scores[idx] if idx < len(entity.scores) else 3
            cell = ws.cell(row=row, column=2 + idx, value=max(1, min(5, int(score))))
            cell.fill = style.input_fill
            cell.alignment = _CENTER
        weighted = ws.cell(
            row=row,
            column=ws_col,
            value=(
                f"=SUMPRODUCT(B{row}:{last_crit_col}{row},"
                f"$B${_WEIGHTS_ROW}:${last_crit_col}${_WEIGHTS_ROW})"
            ),
        )
        weighted.fill = style.formula_fill
        weighted.number_format = "0.00"
        weighted.font = style.bold
        weighted.alignment = _CENTER
        rank = ws.cell(
            row=row,
            column=rank_col,
            value=(
                f"=RANK({ws_letter}{row},"
                f"${ws_letter}${_FIRST_ENTITY_ROW}:${ws_letter}${last_entity_row},0)"
            ),
        )
        rank.fill = style.formula_fill
        rank.alignment = _CENTER

    # Column widths + freeze (header + weights rows + entity-name column stay visible).
    ws.column_dimensions["A"].width = 26
    for idx in range(n):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 13
    ws.column_dimensions[ws_letter].width = 16
    ws.column_dimensions[rank_letter].width = 8
    ws.freeze_panes = f"B{_FIRST_ENTITY_ROW}"

    if data.entities:
        score_range = f"{ws_letter}{_FIRST_ENTITY_ROW}:{ws_letter}{last_entity_row}"
        ws.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color=style.scale_low,
                mid_type="num",
                mid_value=3,
                mid_color=style.scale_mid,
                end_type="num",
                end_value=5,
                end_color=style.scale_high,
            ),
        )
        # Highlight the top-ranked entity's whole row (rank == 1).
        ws.conditional_formatting.add(
            f"A{_FIRST_ENTITY_ROW}:{rank_letter}{last_entity_row}",
            FormulaRule(formula=[f"${rank_letter}{_FIRST_ENTITY_ROW}=1"], fill=style.positive_fill),
        )


def _criteria_guide_sheet(ws: Worksheet, style: _Style, data: DecisionMatrixData) -> None:
    """Write the Criteria_Guide tab: each criterion's weight + 1..5 scoring definition."""
    language = data.language
    _title_block(
        ws,
        style,
        _t(language, "Kriterien & Bewertung", "Criteria & Scoring Guide"),
        _t(
            language,
            "Definition pro Kriterium (1 = schwach … 5 = stark).",
            "Definition per criterion (1 = weak … 5 = strong).",
        ),
    )
    headers = [
        _t(language, "Kriterium", "Criterion"),
        _t(language, "Gewicht", "Weight"),
        _t(language, "Bewertung 1 (schlecht) … 5 (gut)", "How to score 1 (worst) … 5 (best)"),
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER
    weights = _normalized_weights(data)
    for offset, criterion in enumerate(data.criteria):
        row = _HEADER_ROW + 1 + offset
        ws.cell(row=row, column=1, value=criterion.name).font = style.bold
        wcell = ws.cell(
            row=row, column=2, value=round(weights[offset], 4) if offset < len(weights) else None
        )
        wcell.number_format = "0%"
        ws.cell(row=row, column=3, value=criterion.definition or "—").alignment = _LEFT
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = f"A{_HEADER_ROW + 1}"


def _research_notes_sheet(ws: Worksheet, style: _Style, data: DecisionMatrixData) -> None:
    """Write the Research_Notes tab: the evidence behind each entity × criterion score."""
    language = data.language
    _title_block(
        ws,
        style,
        _t(language, "Recherche-Notizen", "Research Notes"),
        _t(language, "Belege/Quellen hinter jedem Score.", "Evidence/sources behind each score."),
    )
    headers = [
        _t(language, "Unternehmen / Option", "Company / Option"),
        _t(language, "Kriterium", "Criterion"),
        _t(language, "Score", "Score"),
        _t(language, "Beleg / Quelle", "Evidence / Source"),
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER
    row = _HEADER_ROW + 1
    for entity in data.entities:
        for idx, criterion in enumerate(data.criteria):
            ws.cell(row=row, column=1, value=entity.name)
            ws.cell(row=row, column=2, value=criterion.name)
            ws.cell(
                row=row, column=3, value=entity.scores[idx] if idx < len(entity.scores) else None
            )
            note = entity.notes[idx] if idx < len(entity.notes) else ""
            ws.cell(row=row, column=4, value=note or "—").alignment = _LEFT
            row += 1
    for col_letter, width in (("A", 24), ("B", 22), ("C", 8), ("D", 70)):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{_HEADER_ROW + 1}"


def build_decision_matrix(
    data: DecisionMatrixData, config: AppConfig, output_dir: str | Path
) -> Path:
    """Build the E-1 Decision/Scoring Matrix workbook and return its path (SPEC §12).

    Weighted scores use ``SUMPRODUCT`` over the yellow weights row; ranks use ``RANK`` — changing
    a weight or a score in Excel recalculates everything (N-10). Three tabs: Summary_Matrix,
    Criteria_Guide, Research_Notes.
    """
    if not data.criteria:
        raise ExcelBuildError("Decision matrix needs at least one scoring criterion.")
    style = _Style(config)
    wb = Workbook()
    matrix = wb.active
    matrix.title = "Summary_Matrix"
    _matrix_sheet(matrix, style, data)
    _criteria_guide_sheet(wb.create_sheet("Criteria_Guide"), style, data)
    _research_notes_sheet(wb.create_sheet("Research_Notes"), style, data)

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / f"{date.today().isoformat()}_{_slug(data.title)}_matrix.xlsx"
    wb.save(str(path))
    return path


# --------------------------------------------------------------- E-2 Intelligence / Benchmark
def _benchmark_table(ws: Worksheet, style: _Style, data: BenchmarkData) -> None:
    """Write the Benchmark_Table tab: an Excel Table (auto-filter, sortable) of entity × metric."""
    language = data.language
    _title_block(
        ws,
        style,
        data.title,
        _t(
            language,
            "Faktenvergleich (keine Bewertung). Spalten per Filter sortierbar.",
            "Factual comparison (no scoring). Sort any column via the filter.",
        ),
    )
    header_row = _HEADER_ROW
    headers = [_t(language, "Unternehmen", "Company"), *data.metrics]
    n_cols = len(headers)
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER
    for offset, row_data in enumerate(data.rows):
        row = header_row + 1 + offset
        name_cell = ws.cell(row=row, column=1, value=row_data.name)
        name_cell.font = style.bold
        if offset % 2:
            name_cell.fill = style.surface_fill
        for idx in range(len(data.metrics)):
            value = row_data.values[idx] if idx < len(row_data.values) else ""
            cell = ws.cell(row=row, column=2 + idx, value=value or "—")
            cell.alignment = _LEFT
            if offset % 2:
                cell.fill = style.surface_fill

    last_row = header_row + len(data.rows)
    ref = f"A{header_row}:{get_column_letter(n_cols)}{max(last_row, header_row + 1)}"
    table = Table(displayName="BenchmarkTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)
    ws.column_dimensions["A"].width = 26
    for idx in range(len(data.metrics)):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 18
    ws.freeze_panes = f"B{header_row + 1}"


def _benchmark_sources(ws: Worksheet, style: _Style, data: BenchmarkData) -> None:
    """Write the Sources tab: per entity × metric provenance (url, date, confidence)."""
    language = data.language
    _title_block(
        ws,
        style,
        _t(language, "Quellen", "Sources"),
        _t(language, "Herkunft jeder Kennzahl.", "Provenance of each value."),
    )
    headers = [
        _t(language, "Unternehmen", "Company"),
        _t(language, "Kennzahl", "Metric"),
        _t(language, "Wert", "Value"),
        _t(language, "Quelle (URL)", "Source URL"),
        _t(language, "Datum", "Date"),
        _t(language, "Konfidenz", "Confidence"),
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER
    for offset, src in enumerate(data.sources):
        row = _HEADER_ROW + 1 + offset
        for col, value in enumerate(
            (src.entity, src.metric, src.value, src.url, src.date, src.confidence), start=1
        ):
            cell = ws.cell(row=row, column=col, value=value or "—")
            cell.alignment = _LEFT
            if src.confidence.lower() == "estimate" and col == 6:
                cell.font = style.muted
    for col_letter, width in (("A", 22), ("B", 20), ("C", 16), ("D", 40), ("E", 12), ("F", 12)):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{_HEADER_ROW + 1}"


def build_benchmark_table(data: BenchmarkData, config: AppConfig, output_dir: str | Path) -> Path:
    """Build the E-2 Intelligence/Benchmark Table workbook and return its path (SPEC §12).

    An Excel Table with auto-filter over the entity × metric grid (sortable by any column), plus a
    Sources tab with per-value provenance + confidence. Facts only — no scoring (that is E-1).
    """
    if not data.metrics:
        raise ExcelBuildError("Benchmark table needs at least one metric column.")
    style = _Style(config)
    wb = Workbook()
    table = wb.active
    table.title = "Benchmark_Table"
    _benchmark_table(table, style, data)
    _benchmark_sources(wb.create_sheet("Sources"), style, data)

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / f"{date.today().isoformat()}_{_slug(data.title)}_benchmark.xlsx"
    wb.save(str(path))
    return path


# --------------------------------------------------------------- E-4 Tracker / Status Dashboard
_STATUS_OPTIONS = ("Active", "On Hold", "Completed", "Dropped")
_PRIORITY_OPTIONS = ("High", "Medium", "Low")
_TRACKER_HEADER_ROW = 3
_TRACKER_FIRST_DATA_ROW = 4
_TRACKER_MAX_ROW = 200  # data-validation + conditional formatting extend over the working range


def _tracker_sheet(ws: Worksheet, style: _Style, data: TrackerData) -> None:
    """Write the Tracker tab: items + Status/Priority dropdowns + conditional formatting."""
    language = data.language
    ws["A1"] = _t(language, "Tracker", "Tracker")
    ws["A1"].font = style.title_font
    headers = [
        _t(language, "Eintrag / Thema", "Entity / Item"),
        _t(language, "Kategorie", "Category"),
        _t(language, "Status", "Status"),
        _t(language, "Priorität", "Priority"),
        _t(language, "Verantwortlich", "Owner"),
        _t(language, "Nächster Schritt", "Next Step"),
        _t(language, "Fällig am", "Next Step Date"),
        _t(language, "Letztes Update", "Last Update"),
        _t(language, "Notizen", "Notes"),
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=_TRACKER_HEADER_ROW, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER
    for offset, item in enumerate(data.items):
        row = _TRACKER_FIRST_DATA_ROW + offset
        values = (
            item.name,
            item.category,
            item.status,
            item.priority,
            item.owner,
            item.next_step,
            item.next_step_date,
            item.last_update,
            item.notes,
        )
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value or "").alignment = _LEFT

    # Data-validation dropdowns (Status col C, Priority col D) over the whole working range.
    status_dv = DataValidation(
        type="list", formula1='"' + ",".join(_STATUS_OPTIONS) + '"', allow_blank=True
    )
    priority_dv = DataValidation(
        type="list", formula1='"' + ",".join(_PRIORITY_OPTIONS) + '"', allow_blank=True
    )
    ws.add_data_validation(status_dv)
    ws.add_data_validation(priority_dv)
    status_dv.add(f"C{_TRACKER_FIRST_DATA_ROW}:C{_TRACKER_MAX_ROW}")
    priority_dv.add(f"D{_TRACKER_FIRST_DATA_ROW}:D{_TRACKER_MAX_ROW}")

    # Conditional formatting: Status colours (Active/On Hold/Dropped) + Priority High = red.
    status_range = f"C{_TRACKER_FIRST_DATA_ROW}:C{_TRACKER_MAX_ROW}"
    ws.conditional_formatting.add(
        status_range, FormulaRule(formula=['$C4="Active"'], fill=style.positive_fill)
    )
    ws.conditional_formatting.add(
        status_range, FormulaRule(formula=['$C4="On Hold"'], fill=style.input_fill)
    )
    ws.conditional_formatting.add(
        status_range, FormulaRule(formula=['$C4="Dropped"'], fill=style.surface_fill)
    )
    ws.conditional_formatting.add(
        f"D{_TRACKER_FIRST_DATA_ROW}:D{_TRACKER_MAX_ROW}",
        FormulaRule(formula=['$D4="High"'], fill=style.negative_fill),
    )
    widths = (26, 16, 12, 10, 16, 30, 13, 14, 34)
    for idx, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + idx)].width = width
    ws.freeze_panes = f"B{_TRACKER_FIRST_DATA_ROW}"


def _dashboard_sheet(ws: Worksheet, style: _Style, language: Language) -> None:
    """Write the Dashboard tab: COUNTIF stats formula-linked to the Tracker tab."""
    ws["A1"] = _t(language, "Dashboard", "Dashboard")
    ws["A1"].font = style.title_font
    rng = f"Tracker!$C${_TRACKER_FIRST_DATA_ROW}:$C${_TRACKER_MAX_ROW}"
    name_rng = f"Tracker!$A${_TRACKER_FIRST_DATA_ROW}:$A${_TRACKER_MAX_ROW}"
    prio_rng = f"Tracker!$D${_TRACKER_FIRST_DATA_ROW}:$D${_TRACKER_MAX_ROW}"
    stats = [
        (_t(language, "Einträge gesamt", "Total items"), f"=COUNTA({name_rng})"),
        (_t(language, "Aktiv", "Active"), f'=COUNTIF({rng},"Active")'),
        (_t(language, "On Hold", "On Hold"), f'=COUNTIF({rng},"On Hold")'),
        (_t(language, "Abgeschlossen", "Completed"), f'=COUNTIF({rng},"Completed")'),
        (_t(language, "Hohe Priorität", "High priority"), f'=COUNTIF({prio_rng},"High")'),
    ]
    for offset, (label, formula) in enumerate(stats):
        row = _TRACKER_HEADER_ROW + offset
        ws.cell(row=row, column=1, value=label).font = style.bold
        cell = ws.cell(row=row, column=2, value=formula)
        cell.fill = style.formula_fill
        cell.font = style.bold
        cell.alignment = _CENTER
    updated_row = _TRACKER_HEADER_ROW + len(stats) + 1
    ws.cell(
        row=updated_row, column=1, value=_t(language, "Stand", "Last updated")
    ).font = style.bold
    upd = ws.cell(row=updated_row, column=2)
    upd.fill = style.input_fill  # yellow: user fills the date
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16


def _archive_sheet(ws: Worksheet, style: _Style, data: TrackerData) -> None:
    """Write the Archive tab: same headers as Tracker, for completed/dropped items (user-kept)."""
    _tracker_sheet(ws, style, TrackerData(title=data.title, language=data.language, items=[]))
    ws["A1"] = _t(data.language, "Archiv", "Archive")
    ws["A1"].font = style.title_font


def build_tracker(data: TrackerData, config: AppConfig, output_dir: str | Path) -> Path:
    """Build the E-4 Tracker / Status Dashboard workbook and return its path (SPEC §12).

    Three tabs: Dashboard (COUNTIF stats formula-linked to Tracker), Tracker (Status/Priority
    data-validation dropdowns + conditional formatting), Archive (same shape, user-managed). The
    dropdowns + colour-coding make this the one living document the user maintains.
    """
    style = _Style(config)
    wb = Workbook()
    dashboard = wb.active
    dashboard.title = "Dashboard"
    _dashboard_sheet(dashboard, style, data.language)
    _tracker_sheet(wb.create_sheet("Tracker"), style, data)
    _archive_sheet(wb.create_sheet("Archive"), style, data)

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / f"{date.today().isoformat()}_{_slug(data.title)}_tracker.xlsx"
    wb.save(str(path))
    return path


# --------------------------------------------------------------- E-3 Business Case Model
# Assumptions tab layout: each driver lives in a named cell so every other tab references it.
# Rows (1-based) on the Assumptions sheet:
_A = {
    "investment": 4,
    "revenue_year1": 5,
    "revenue_growth": 6,
    "opex_year1": 7,
    "opex_growth": 8,
    "discount_rate": 9,
}
_ASSUMPTION_VALUE_COL = "B"  # yellow input column on the Assumptions tab
_EXTRA_ASSUMPTION_ROW = 11  # extra labelled assumptions start here


def _aref(key: str) -> str:
    """Absolute reference to a core assumption's yellow input cell (e.g. ``Assumptions!$B$5``)."""
    return f"Assumptions!${_ASSUMPTION_VALUE_COL}${_A[key]}"


def _assumptions_sheet(ws: Worksheet, style: _Style, data: BusinessCaseData) -> None:
    """Write the Assumptions tab — ALL yellow inputs; every other tab references these (N-10)."""
    language = data.language
    _title_block(
        ws,
        style,
        _t(language, "Annahmen", "Assumptions"),
        _t(
            language,
            "NUR gelbe Zellen bearbeiten — alle anderen Tabs rechnen daraus neu.",
            "Edit ONLY the yellow cells — every other tab recalculates from these.",
        ),
    )
    rows = [
        ("investment", _t(language, "Einmalinvestition", "One-time investment"), "EUR"),
        ("revenue_year1", _t(language, "Umsatz Jahr 1", "Revenue Year 1"), "EUR"),
        ("revenue_growth", _t(language, "Umsatzwachstum p.a.", "Revenue growth p.a."), "%"),
        ("opex_year1", _t(language, "OpEx Jahr 1", "OpEx Year 1"), "EUR"),
        ("opex_growth", _t(language, "OpEx-Wachstum p.a.", "OpEx growth p.a."), "%"),
        ("discount_rate", _t(language, "Diskontsatz (NPV)", "Discount rate (NPV)"), "%"),
    ]
    values = {
        "investment": data.investment,
        "revenue_year1": data.revenue_year1,
        "revenue_growth": data.revenue_growth,
        "opex_year1": data.opex_year1,
        "opex_growth": data.opex_growth,
        "discount_rate": data.discount_rate,
    }
    for key, label, unit in rows:
        row = _A[key]
        ws.cell(row=row, column=1, value=label).font = style.bold
        cell = ws.cell(row=row, column=2, value=values[key])
        cell.fill = style.input_fill
        cell.number_format = "0%" if unit == "%" else "#,##0"
        ws.cell(row=row, column=3, value=unit).font = style.muted

    # Extra labelled assumptions (also yellow), surfaced for the audit trail.
    for offset, assumption in enumerate(data.assumptions):
        row = _EXTRA_ASSUMPTION_ROW + offset
        ws.cell(row=row, column=1, value=assumption.name).font = style.bold
        cell = ws.cell(row=row, column=2, value=assumption.value)
        cell.fill = style.input_fill
        cell.number_format = "0%" if assumption.unit == "%" else "#,##0"
        ws.cell(row=row, column=3, value=assumption.unit).font = style.muted
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10


def _projections_sheet(ws: Worksheet, style: _Style, data: BusinessCaseData) -> None:
    """Write the Projections tab: annual Revenue/OpEx/EBITDA/Cash Flow — all formula-linked."""
    language = data.language
    _title_block(
        ws,
        style,
        _t(language, "Projektionen", "Financial Projections"),
        _t(language, "Formelbasiert aus 'Annahmen'.", "Formula-driven from 'Assumptions'."),
    )
    header_row = 5
    labels = [
        _t(language, "Jahr", "Year"),
        _t(language, "Umsatz", "Revenue"),
        _t(language, "OpEx", "OpEx"),
        _t(language, "EBITDA", "EBITDA"),
        _t(language, "Cashflow", "Cash Flow"),
        _t(language, "Kumulierter Cashflow", "Cumulative Cash Flow"),
    ]
    for col, text in enumerate(labels, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER

    # Year 0 = the upfront investment (negative cash flow), then N project years.
    year0 = header_row + 1
    ws.cell(row=year0, column=1, value=0).alignment = _CENTER
    for col in (2, 3, 4):
        ws.cell(row=year0, column=col, value=0)
    cf0 = ws.cell(row=year0, column=5, value=f"=-{_aref('investment')}")
    cf0.fill = style.formula_fill
    cf0.number_format = "#,##0"
    cum0 = ws.cell(row=year0, column=6, value="=E6")
    cum0.fill = style.formula_fill
    cum0.number_format = "#,##0"

    for yr in range(1, data.years + 1):
        row = year0 + yr
        ws.cell(row=row, column=1, value=yr).alignment = _CENTER
        if yr == 1:
            revenue = f"={_aref('revenue_year1')}"
            opex = f"={_aref('opex_year1')}"
        else:
            prev = row - 1
            revenue = f"=B{prev}*(1+{_aref('revenue_growth')})"
            opex = f"=C{prev}*(1+{_aref('opex_growth')})"
        for col, formula in ((2, revenue), (3, opex)):
            cell = ws.cell(row=row, column=col, value=formula)
            cell.fill = style.formula_fill
            cell.number_format = "#,##0"
        ebitda = ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
        cashflow = ws.cell(row=row, column=5, value=f"=D{row}")
        cumulative = ws.cell(row=row, column=6, value=f"=F{row - 1}+E{row}")
        for cell in (ebitda, cashflow, cumulative):
            cell.fill = style.formula_fill
            cell.number_format = "#,##0"
        # Colour cumulative cash flow green when positive, red when negative (breakeven visual).
        ws.conditional_formatting.add(
            f"F{row}", FormulaRule(formula=[f"F{row}>=0"], fill=style.positive_fill)
        )
        ws.conditional_formatting.add(
            f"F{row}", FormulaRule(formula=[f"F{row}<0"], fill=style.negative_fill)
        )
    for col_letter in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 18
    ws.freeze_panes = "B6"


def _last_projection_row(data: BusinessCaseData) -> int:
    """Row index of the final project year on the Projections tab."""
    return 6 + data.years  # header_row(5) + year0(6) is row 6; +years more rows


def _summary_sheet(ws: Worksheet, style: _Style, data: BusinessCaseData) -> None:
    """Write the Executive Summary tab: NPV / IRR / payback — all formulas over the projections."""
    language = data.language
    _title_block(
        ws,
        style,
        _t(language, "Zusammenfassung", "Executive Summary"),
        _t(language, "Kennzahlen — formelbasiert.", "Headline metrics — all formula-driven."),
    )
    last = _last_projection_row(data)
    cf_range = f"Projections!E7:E{last}"  # year 1..N cash flows (year 0 is the investment)
    metrics = [
        (
            _t(language, "Gesamtinvestition", "Total Investment"),
            f"={_aref('investment')}",
            "#,##0",
        ),
        (
            _t(language, "NPV (3-5 J.)", "NPV"),
            f"=-{_aref('investment')}+NPV({_aref('discount_rate')},{cf_range})",
            "#,##0",
        ),
        (
            _t(language, "IRR", "IRR"),
            f"=IFERROR(IRR(Projections!E6:E{last}),0)",
            "0.0%",
        ),
        (
            _t(language, "Kumulierter Cashflow (Endjahr)", "Cumulative Cash Flow (final year)"),
            f"=Projections!F{last}",
            "#,##0",
        ),
    ]
    for offset, (label, formula, fmt) in enumerate(metrics):
        row = 5 + offset
        ws.cell(row=row, column=1, value=label).font = style.bold
        cell = ws.cell(row=row, column=2, value=formula)
        cell.fill = style.formula_fill
        cell.font = style.bold
        cell.number_format = fmt
    bl_row = 5 + len(metrics) + 1
    ws.cell(
        row=bl_row, column=1, value=_t(language, "Kernaussage", "Bottom Line")
    ).font = style.bold
    ws.cell(row=bl_row, column=2, value=data.bottom_line or "—").alignment = _LEFT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40


def _scenarios_sheet(ws: Worksheet, style: _Style, data: BusinessCaseData) -> None:
    """Write the Scenarios tab: Base/Optimistic/Pessimistic NPV via yellow multiplier inputs."""
    language = data.language
    _title_block(
        ws,
        style,
        _t(language, "Szenarien", "Scenarios"),
        _t(
            language,
            "Gelbe Multiplikatoren anpassen — NPV je Szenario rechnet neu.",
            "Adjust the yellow multipliers — each scenario's NPV recalculates.",
        ),
    )
    header_row = 5
    headers = [
        _t(language, "Szenario", "Scenario"),
        _t(language, "Umsatz-Faktor", "Revenue factor"),
        _t(language, "Kosten-Faktor", "Cost factor"),
        "NPV",
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER
    last = _last_projection_row(data)
    # NPV per scenario: scale year 1..N revenue and opex by the (yellow) factors.
    # cash flow_yr = rev_yr*revfactor - opex_yr*costfactor; discounted via NPV().
    scenarios = [
        (_t(language, "Base", "Base Case"), 1.0, 1.0),
        (_t(language, "Optimistisch", "Optimistic"), 1.2, 0.9),
        (_t(language, "Pessimistisch", "Pessimistic"), 0.8, 1.15),
    ]
    for offset, (label, rev_factor, cost_factor) in enumerate(scenarios):
        row = header_row + 1 + offset
        ws.cell(row=row, column=1, value=label).font = style.bold
        rev_cell = ws.cell(row=row, column=2, value=rev_factor)
        cost_cell = ws.cell(row=row, column=3, value=cost_factor)
        for cell in (rev_cell, cost_cell):
            cell.fill = style.input_fill
            cell.number_format = "0%"
            cell.alignment = _CENTER
        # SUMPRODUCT of discounted scaled cash flows minus the upfront investment.
        rev_range = f"Projections!B7:B{last}"
        opex_range = f"Projections!C7:C{last}"
        npv_formula = (
            f"=-{_aref('investment')}+NPV({_aref('discount_rate')},"
            f"({rev_range}*$B${row})-({opex_range}*$C${row}))"
        )
        npv = ws.cell(row=row, column=4, value=npv_formula)
        npv.fill = style.formula_fill
        npv.number_format = "#,##0"
        npv.font = style.bold
        ws.conditional_formatting.add(
            f"D{row}", FormulaRule(formula=[f"D{row}>=0"], fill=style.positive_fill)
        )
        ws.conditional_formatting.add(
            f"D{row}", FormulaRule(formula=[f"D{row}<0"], fill=style.negative_fill)
        )
    for col_letter, width in (("A", 18), ("B", 14), ("C", 14), ("D", 18)):
        ws.column_dimensions[col_letter].width = width


def _audit_sheet(ws: Worksheet, style: _Style, data: BusinessCaseData) -> None:
    """Write the Sources & Audit Trail tab: every assumption → its value, source, confidence."""
    language = data.language
    _title_block(
        ws,
        style,
        _t(language, "Quellen & Audit-Trail", "Sources & Audit Trail"),
        _t(language, "Herkunft jeder Annahme.", "Provenance of every assumption."),
    )
    headers = [
        _t(language, "Annahme", "Assumption"),
        _t(language, "Wert", "Value"),
        _t(language, "Einheit", "Unit"),
        _t(language, "Quelle", "Source"),
        _t(language, "Konfidenz", "Confidence"),
    ]
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=text)
        cell.fill = style.header_fill
        cell.font = style.header_font
        cell.alignment = _CENTER
    for offset, assumption in enumerate(data.assumptions):
        row = _HEADER_ROW + 1 + offset
        ws.cell(row=row, column=1, value=assumption.name)
        vcell = ws.cell(row=row, column=2, value=assumption.value)
        vcell.number_format = "0%" if assumption.unit == "%" else "#,##0"
        ws.cell(row=row, column=3, value=assumption.unit or "—")
        ws.cell(row=row, column=4, value=assumption.source or "—").alignment = _LEFT
        conf = ws.cell(row=row, column=5, value=assumption.confidence or "—")
        if assumption.confidence.lower() == "estimate":
            conf.font = style.muted
    for col_letter, width in (("A", 28), ("B", 14), ("C", 10), ("D", 44), ("E", 12)):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{_HEADER_ROW + 1}"


def build_business_case(data: BusinessCaseData, config: AppConfig, output_dir: str | Path) -> Path:
    """Build the E-3 Business Case financial model and return its path (SPEC §12, N-10).

    Five formula-linked tabs: Executive Summary (NPV/IRR/payback formulas), Assumptions (ALL yellow
    inputs), Financial Projections (Revenue/OpEx/EBITDA/Cash Flow — every cell references the
    Assumptions), Scenarios (Base/Optimistic/Pessimistic NPV via yellow multipliers), Sources &
    Audit Trail. Changing one yellow assumption recalculates the entire model in Excel.
    """
    style = _Style(config)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _assumptions_sheet(wb.create_sheet("Assumptions"), style, data)
    _projections_sheet(wb.create_sheet("Projections"), style, data)
    _scenarios_sheet(wb.create_sheet("Scenarios"), style, data)
    _audit_sheet(wb.create_sheet("Sources"), style, data)
    # Summary references Projections/Assumptions, so fill it after the others exist.
    _summary_sheet(summary, style, data)

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / f"{date.today().isoformat()}_{_slug(data.title)}_business_case.xlsx"
    wb.save(str(path))
    return path


# --------------------------------------------------------------- dispatcher
def build_workbook(
    template: ExcelTemplate,
    data: DecisionMatrixData | BenchmarkData | BusinessCaseData | TrackerData,
    config: AppConfig,
    output_dir: str | Path,
) -> Path:
    """Render the shaped ``(template, data)`` pair into its .xlsx (dispatches to E-1..E-4).

    The ``data`` type matches ``template`` (both come from ``content_shaper.shape_workbook``);
    a mismatch is a programming error and raises :class:`ExcelBuildError`.
    """
    if template == ExcelTemplate.BENCHMARK_TABLE and isinstance(data, BenchmarkData):
        return build_benchmark_table(data, config, output_dir)
    if template == ExcelTemplate.BUSINESS_CASE_MODEL and isinstance(data, BusinessCaseData):
        return build_business_case(data, config, output_dir)
    if template == ExcelTemplate.TRACKER_DASHBOARD and isinstance(data, TrackerData):
        return build_tracker(data, config, output_dir)
    if template == ExcelTemplate.DECISION_MATRIX and isinstance(data, DecisionMatrixData):
        return build_decision_matrix(data, config, output_dir)
    raise ExcelBuildError(f"Workbook data type {type(data).__name__} does not match {template}.")
