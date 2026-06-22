"""Analyst dimension — Recruiting CV screening & ranking (Dimensions / Phase 6).

Local, DSGVO-safe screening: read each CV, score it against weighted job criteria with the local
LLM under the zero-hallucination rubric (``playbooks/recruiting_screening_playbook.md`` — only what
the CV states; evidence required), then rank and emit an Excel ranking workbook. **The human
recruiter decides** — this is a defensible shortlist, not a hiring decision.

All LLM access goes through :class:`LocalLLMClient` (RULE 6). Parsing is tolerant
(:func:`core.json_utils.extract_json_object`) and **fails open**: an unparsable evaluation becomes
a zero-scored candidate with an explicit gap, never a crash.
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

from core.config import AppConfig
from core.json_utils import extract_json_object
from llm.local_llm_client import LocalLLMClient
from models.research import DocContent
from models.scoring import (
    CandidateEvaluation,
    CriterionScore,
    JobCriterion,
    ScreeningResult,
)

_PLAYBOOK_FILE = (
    Path(__file__).resolve().parent.parent / "playbooks" / "recruiting_screening_playbook.md"
)
_MAX_CV_CHARS = 12000  # cap CV text per scoring call (safe for a 32K+ context budget)
_MAX_JOB_CHARS = 8000

_FALLBACK_RUBRIC = (
    "Score each candidate ONLY on what their CV states. Require a short verbatim quote as evidence "
    "for any non-zero score; if a criterion is not supported, score it low and leave evidence "
    "empty. Never invent experience. Ignore name, gender, age, and nationality. The human "
    "recruiter makes the final decision."
)

DEFAULT_CRITERIA: list[JobCriterion] = [
    JobCriterion(name="Relevant experience", weight=3.0, description="Depth in a comparable role"),
    JobCriterion(
        name="Required skills", weight=3.0, description="Hard skills/tools the role needs"
    ),
    JobCriterion(name="Domain fit", weight=2.0, description="Industry/domain relevance"),
    JobCriterion(
        name="Education / qualifications", weight=1.0, description="Degrees, certifications"
    ),
    JobCriterion(
        name="Impact / achievements", weight=1.0, description="Quantified results, ownership"
    ),
]


# --------------------------------------------------------------- helpers
def _slug(text: str) -> str:
    """Make a short, filesystem-safe slug from a title."""
    cleaned = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return (cleaned or "role")[:40]


def _argb(hex_color: str) -> str:
    """Normalize a ``#rrggbb`` color to openpyxl 8-digit ARGB."""
    value = hex_color.lstrip("#").upper()
    return value if len(value) == 8 else "FF" + value


def _clamp(value: object, low: float, high: float, default: float) -> float:
    """Coerce ``value`` to float and clamp to [low, high]; ``default`` on failure."""
    try:
        number = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return default
    return max(low, min(high, number))


def _load_rubric() -> str:
    """Load the recruiting playbook text (best-effort; fallback rubric if missing)."""
    try:
        text = _PLAYBOOK_FILE.read_text(encoding="utf-8").strip()
    except OSError:
        return _FALLBACK_RUBRIC
    return text or _FALLBACK_RUBRIC


def parse_criteria_arg(text: str) -> list[JobCriterion]:
    """Parse a ``"a, b, c"`` criteria string into equally-weighted :class:`JobCriterion`."""
    names = [part.strip() for part in text.split(",") if part.strip()]
    return [JobCriterion(name=name, weight=1.0) for name in names]


def compute_weighted_score(scores: list[CriterionScore], criteria: list[JobCriterion]) -> float:
    """Weighted average (0-100) of criterion scores using the criteria weights."""
    weight_by_name = {c.name.strip().lower(): c.weight for c in criteria}
    accumulated = 0.0
    total_weight = 0.0
    seen: set[str] = set()
    for entry in scores:
        key = entry.criterion.strip().lower()
        weight = weight_by_name.get(key)
        if weight is None or key in seen:
            continue
        seen.add(key)
        accumulated += entry.score * weight
        total_weight += weight
    if total_weight <= 0.0:
        return 0.0
    return round(accumulated / total_weight, 1)


# --------------------------------------------------------------- LLM steps
def derive_criteria(
    llm: LocalLLMClient, job_profile: str, *, max_criteria: int = 6
) -> tuple[str, list[JobCriterion]]:
    """Extract a job title + weighted criteria from the profile (fallback to defaults)."""
    prompt = (
        f"From this job profile, extract the {max_criteria} most important weighted hiring "
        f'criteria.\nJOB PROFILE:\n"""{job_profile[:_MAX_JOB_CHARS]}"""\n\n'
        'Return ONLY JSON: {"job_title": "<title>", "criteria": '
        '[{"name": "<short criterion>", "weight": <1-5>, "description": "<what good looks like>"}]}'
    )
    raw = llm.generate(
        prompt, system="You extract concise, weighted hiring criteria. Output only JSON."
    )
    data = extract_json_object(raw)
    if not data:
        return "", DEFAULT_CRITERIA
    job_title = str(data.get("job_title", "")).strip()
    criteria: list[JobCriterion] = []
    for item in data.get("criteria", []) or []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()
        if not name:
            continue
        criteria.append(
            JobCriterion(
                name=name,
                weight=_clamp(item.get("weight", 1.0), 0.1, 5.0, 1.0),
                description=str(item.get("description", "")).strip(),
            )
        )
    if not criteria:
        return job_title, DEFAULT_CRITERIA
    return job_title, criteria


def score_candidate(
    llm: LocalLLMClient,
    source_file: str,
    cv_text: str,
    criteria: list[JobCriterion],
    *,
    rubric: str | None = None,
) -> CandidateEvaluation:
    """Score one CV against the criteria and return a :class:`CandidateEvaluation` (fail-open)."""
    rubric_text = rubric or _load_rubric()
    criteria_lines = "\n".join(
        f"- {c.name} (weight {c.weight:g}): {c.description}" for c in criteria
    )
    prompt = (
        f"CRITERIA (score each 0-100):\n{criteria_lines}\n\n"
        f'CANDIDATE CV (verbatim):\n"""{cv_text[:_MAX_CV_CHARS]}"""\n\n'
        "Return ONLY JSON:\n"
        '{"candidate": "<full name from the CV or \'Unknown\'>", "criterion_scores": '
        '[{"criterion": "<exact criterion name>", "score": <0-100>, "rationale": "<one sentence>", '
        '"evidence": "<short verbatim quote from the CV, or empty>"}], '
        '"summary": "<2-3 sentence fit assessment>", "strengths": ["..."], '
        '"gaps": ["<requirement not evidenced in the CV>"]}'
    )
    system = (
        f"{rubric_text}\n\nYou are screening ONE CV against the criteria. "
        "Output only a single JSON object."
    )
    raw = llm.generate(prompt, system=system)
    data = extract_json_object(raw)
    stem = Path(source_file).stem if source_file else "candidate"
    if not data:
        return CandidateEvaluation(
            candidate=stem,
            source_file=source_file,
            summary="Could not parse the model's evaluation.",
            gaps=["Evaluation could not be parsed — re-run or review this CV manually."],
        )
    scores: list[CriterionScore] = []
    for item in data.get("criterion_scores", []) or []:
        if not isinstance(item, dict):
            continue
        criterion_name = str(item.get("criterion", "")).strip()
        if not criterion_name:
            continue
        scores.append(
            CriterionScore(
                criterion=criterion_name,
                score=_clamp(item.get("score", 0), 0.0, 100.0, 0.0),
                rationale=str(item.get("rationale", "")).strip(),
                evidence=str(item.get("evidence", "")).strip(),
            )
        )
    return CandidateEvaluation(
        candidate=str(data.get("candidate", "")).strip() or stem,
        source_file=source_file,
        criterion_scores=scores,
        weighted_score=compute_weighted_score(scores, criteria),
        summary=str(data.get("summary", "")).strip(),
        strengths=[str(x).strip() for x in (data.get("strengths") or []) if str(x).strip()],
        gaps=[str(x).strip() for x in (data.get("gaps") or []) if str(x).strip()],
    )


def screen_cvs(
    llm: LocalLLMClient,
    job_profile: str,
    cvs: list[DocContent],
    *,
    criteria: list[JobCriterion] | None = None,
    language: str = "en",
) -> ScreeningResult:
    """Screen CVs against a job profile: derive/accept criteria, score each, rank best-first."""
    job_title = ""
    if criteria is None:
        job_title, criteria = derive_criteria(llm, job_profile)
    rubric = _load_rubric()
    evaluations = [
        score_candidate(llm, str(cv.source_path), cv.text, criteria, rubric=rubric) for cv in cvs
    ]
    evaluations.sort(key=lambda evaluation: evaluation.weighted_score, reverse=True)
    for index, evaluation in enumerate(evaluations, start=1):
        evaluation.rank = index
    return ScreeningResult(
        job_title=job_title, criteria=criteria, candidates=evaluations, language=language
    )


# --------------------------------------------------------------- rendering
def render_screening_markdown(result: ScreeningResult) -> str:
    """Render a ranked screening result as a Markdown table (for the console)."""
    lines = [f"# CV Screening — {result.job_title or 'Role'}", ""]
    if result.criteria:
        lines.append("Criteria: " + ", ".join(f"{c.name} (w{c.weight:g})" for c in result.criteria))
        lines.append("")
    lines.append("| Rank | Candidate | Score | Summary |")
    lines.append("|---:|---|---:|---|")
    for evaluation in result.candidates:
        summary = evaluation.summary[:120].replace("\n", " ")
        lines.append(
            f"| {evaluation.rank} | {evaluation.candidate} | "
            f"{evaluation.weighted_score:g} | {summary} |"
        )
    return "\n".join(lines)


def render_screening_excel(result: ScreeningResult, out_dir: Path | str, config: AppConfig) -> Path:
    """Write a Neura-styled Excel ranking workbook and return its path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    colors = config.output.colors
    header_fill = PatternFill("solid", fgColor=_argb(colors.excel_header))
    header_font = Font(bold=True, color=_argb(colors.white))
    positive_fill = PatternFill("solid", fgColor=_argb(colors.excel_positive))
    negative_fill = PatternFill("solid", fgColor=_argb(colors.excel_negative))
    wrap_top = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CV Ranking"
    worksheet["A1"] = f"CV Screening — {result.job_title or 'Role'}"
    worksheet["A1"].font = Font(bold=True, size=14, color=_argb(colors.text_dark))
    worksheet["A2"] = (
        f"Generated {date.today().isoformat()} · Local & DSGVO-safe · "
        "The human recruiter decides. Every score is traceable to the CV."
    )
    worksheet["A2"].font = Font(italic=True, size=9, color=_argb(colors.charcoal))

    headers = (
        ["Rank", "Candidate", "Source"]
        + [c.name for c in result.criteria]
        + ["Weighted Score", "Summary", "Gaps"]
    )
    header_row = 4
    for column, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    score_column = 3 + len(result.criteria) + 1
    row = header_row
    for evaluation in result.candidates:
        row += 1
        score_by_name = {s.criterion.strip().lower(): s.score for s in evaluation.criterion_scores}
        values: list[object] = [
            evaluation.rank,
            evaluation.candidate,
            Path(evaluation.source_file).name if evaluation.source_file else "",
        ]
        values += [score_by_name.get(c.name.strip().lower(), 0.0) for c in result.criteria]
        values += [evaluation.weighted_score, evaluation.summary, "; ".join(evaluation.gaps)]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=column, value=value)
            cell.alignment = wrap_top
        score_cell = worksheet.cell(row=row, column=score_column)
        score_cell.fill = positive_fill if evaluation.weighted_score >= 60 else negative_fill

    worksheet.column_dimensions["A"].width = 6
    worksheet.column_dimensions["B"].width = 22
    worksheet.column_dimensions["C"].width = 18
    for index in range(len(result.criteria)):
        worksheet.column_dimensions[get_column_letter(4 + index)].width = 12
    worksheet.column_dimensions[get_column_letter(score_column)].width = 14
    worksheet.column_dimensions[get_column_letter(score_column + 1)].width = 50
    worksheet.column_dimensions[get_column_letter(score_column + 2)].width = 30

    output_dir = Path(out_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{date.today().isoformat()}_{_slug(result.job_title or 'role')}_cv_ranking.xlsx"
    path = output_dir / filename
    workbook.save(str(path))
    return path
