"""Tests for the Builder dimension — Finance management reporting (Dimensions / Phase 6).

A StubLLM returns canned JSON so consolidation/rendering is tested deterministically with no real
model. Covers source-traced key figures, fail-open parsing, and the Markdown + Excel output.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from core.config import AppConfig
from core.finance_reporting import build_report, render_report_excel, render_report_markdown
from models.research import DocContent


class StubLLM:
    """Minimal LocalLLMClient stand-in: returns queued responses, records prompts."""

    def __init__(self, responses: list[str]) -> None:
        self._responses = list(responses)
        self.prompts: list[str] = []

    def generate(self, prompt: str, system: str = "", **kwargs: Any) -> str:
        self.prompts.append(prompt)
        return self._responses.pop(0) if self._responses else "{}"


def _doc(name: str, text: str = "figures") -> DocContent:
    return DocContent(source_path=Path(name), doc_type="xlsx", text=text)


def test_build_report_parses_key_figures_with_source() -> None:
    payload = {
        "title": "Q2 Report",
        "period": "Q2 2026",
        "bottom_line": "On track",
        "key_figures": [
            {"metric": "Revenue", "value": "€4.2M", "period": "Q2 2026", "source": "q2.xlsx!Sheet1"}
        ],
        "sections": [{"heading": "Revenue up", "bullets": ["b1"]}],
        "risks": ["r1"],
        "gaps": ["g1"],
    }
    llm = StubLLM([json.dumps(payload)])
    report = build_report(llm, [_doc("q2.xlsx", "Revenue 4.2M")], period="Q2 2026")  # type: ignore[arg-type]
    assert report.title == "Q2 Report"
    assert report.key_figures[0].metric == "Revenue"
    assert report.key_figures[0].value == "€4.2M"
    assert report.key_figures[0].source == "q2.xlsx!Sheet1"
    assert report.sources == ["q2.xlsx"]
    assert report.risks == ["r1"]


def test_build_report_failopen_on_garbage() -> None:
    llm = StubLLM(["no json here"])
    report = build_report(llm, [_doc("a.pdf", "x")], title="My Report")  # type: ignore[arg-type]
    assert report.title == "My Report"
    assert report.gaps  # explicit parse-failure gap
    assert report.sources == ["a.pdf"]


def test_render_report_markdown_contains_sections() -> None:
    payload = {
        "title": "R",
        "period": "Q2",
        "bottom_line": "BL",
        "key_figures": [{"metric": "Rev", "value": "€1M", "period": "Q2", "source": "s"}],
        "sections": [{"heading": "Growth", "bullets": ["x"]}],
        "risks": ["risk"],
        "gaps": ["gap"],
    }
    llm = StubLLM([json.dumps(payload)])
    report = build_report(llm, [_doc("s.xlsx")])  # type: ignore[arg-type]
    markdown = render_report_markdown(report)
    assert "Bottom Line" in markdown and "BL" in markdown
    assert "Key Figures" in markdown and "Rev" in markdown and "€1M" in markdown
    assert "## Growth" in markdown


def test_render_report_excel_writes_readable_file(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    payload = {
        "title": "R",
        "key_figures": [
            {"metric": "Revenue", "value": "€4.2M", "period": "Q2", "source": "q2.xlsx"}
        ],
    }
    llm = StubLLM([json.dumps(payload)])
    report = build_report(llm, [_doc("q2.xlsx")])  # type: ignore[arg-type]
    out = render_report_excel(report, tmp_path, AppConfig())

    assert out.is_file()
    worksheet = load_workbook(out).active
    assert any(cell.value == "Revenue" for row in worksheet.iter_rows() for cell in row)
    assert any(cell.value == "q2.xlsx" for row in worksheet.iter_rows() for cell in row)
