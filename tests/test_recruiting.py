"""Tests for the Analyst dimension — Recruiting CV screening (Dimensions / Phase 6).

A StubLLM returns canned JSON so scoring/ranking/rendering is tested deterministically with no
real model. Covers weighting, fail-open parsing, ranking order, criteria derivation, and the
Excel ranking output (reopened with openpyxl).
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from core.config import AppConfig
from core.recruiting import (
    DEFAULT_CRITERIA,
    compute_weighted_score,
    derive_criteria,
    parse_criteria_arg,
    render_screening_excel,
    render_screening_markdown,
    score_candidate,
    screen_cvs,
)
from models.research import DocContent
from models.scoring import CriterionScore, JobCriterion


class StubLLM:
    """Minimal LocalLLMClient stand-in: returns queued responses, records prompts."""

    def __init__(self, responses: list[str]) -> None:
        self._responses = list(responses)
        self.prompts: list[str] = []

    def generate(self, prompt: str, system: str = "", **kwargs: Any) -> str:
        self.prompts.append(prompt)
        return self._responses.pop(0) if self._responses else "{}"


def _cv(name: str, text: str = "cv text") -> DocContent:
    return DocContent(source_path=Path(name), doc_type="pdf", text=text)


def test_parse_criteria_arg() -> None:
    crits = parse_criteria_arg("Python, Leadership ,  ML ")
    assert [c.name for c in crits] == ["Python", "Leadership", "ML"]
    assert all(c.weight == 1.0 for c in crits)


def test_compute_weighted_score() -> None:
    criteria = [JobCriterion(name="A", weight=3.0), JobCriterion(name="B", weight=1.0)]
    scores = [CriterionScore(criterion="A", score=80.0), CriterionScore(criterion="B", score=40.0)]
    assert compute_weighted_score(scores, criteria) == 70.0  # (80*3 + 40*1) / 4


def test_compute_weighted_score_ignores_unknown_and_empty() -> None:
    criteria = [JobCriterion(name="A", weight=2.0)]
    scores = [CriterionScore(criterion="A", score=50.0), CriterionScore(criterion="X", score=100.0)]
    assert compute_weighted_score(scores, criteria) == 50.0
    assert compute_weighted_score([], criteria) == 0.0


def test_score_candidate_parses_and_weights() -> None:
    payload = {
        "candidate": "Jane Doe",
        "criterion_scores": [
            {"criterion": "A", "score": 90, "rationale": "r", "evidence": "e"},
            {"criterion": "B", "score": 50, "rationale": "", "evidence": ""},
        ],
        "summary": "Strong fit",
        "strengths": ["x"],
        "gaps": ["y"],
    }
    llm = StubLLM([json.dumps(payload)])
    criteria = [JobCriterion(name="A", weight=3.0), JobCriterion(name="B", weight=1.0)]
    evaluation = score_candidate(llm, "cv1.docx", "cv text", criteria)  # type: ignore[arg-type]
    assert evaluation.candidate == "Jane Doe"
    assert evaluation.weighted_score == 80.0  # (90*3 + 50) / 4
    assert evaluation.summary == "Strong fit"
    assert evaluation.gaps == ["y"]


def test_score_candidate_clamps_out_of_range() -> None:
    payload = {"candidate": "X", "criterion_scores": [{"criterion": "A", "score": 150}]}
    llm = StubLLM([json.dumps(payload)])
    evaluation = score_candidate(llm, "x.pdf", "t", [JobCriterion(name="A", weight=1.0)])  # type: ignore[arg-type]
    assert evaluation.criterion_scores[0].score == 100.0


def test_score_candidate_failopen_on_garbage() -> None:
    llm = StubLLM(["sorry, there is no json here"])
    evaluation = score_candidate(llm, "weird.pdf", "t", [JobCriterion(name="A", weight=1.0)])  # type: ignore[arg-type]
    assert evaluation.candidate == "weird"  # filename-stem fallback
    assert evaluation.weighted_score == 0.0
    assert evaluation.gaps  # explicit "could not parse" gap


def test_derive_criteria_fallback_on_garbage() -> None:
    llm = StubLLM(["not json"])
    _title, criteria = derive_criteria(llm, "some profile")  # type: ignore[arg-type]
    assert criteria == DEFAULT_CRITERIA


def test_derive_criteria_parses() -> None:
    payload = {
        "job_title": "ML Engineer",
        "criteria": [{"name": "Python", "weight": 4, "description": "d"}],
    }
    llm = StubLLM([json.dumps(payload)])
    title, criteria = derive_criteria(llm, "profile")  # type: ignore[arg-type]
    assert title == "ML Engineer"
    assert criteria[0].name == "Python"
    assert criteria[0].weight == 4.0


def test_screen_cvs_ranks_descending() -> None:
    low = {"candidate": "Low", "criterion_scores": [{"criterion": "A", "score": 30}]}
    high = {"candidate": "High", "criterion_scores": [{"criterion": "A", "score": 90}]}
    llm = StubLLM([json.dumps(low), json.dumps(high)])  # cv order: a→low, b→high
    cvs = [_cv("a.pdf"), _cv("b.pdf")]
    result = screen_cvs(llm, "profile", cvs, criteria=[JobCriterion(name="A", weight=1.0)])  # type: ignore[arg-type]
    assert [c.candidate for c in result.candidates] == ["High", "Low"]
    assert result.candidates[0].rank == 1
    assert result.candidates[1].rank == 2


def test_render_markdown_lists_ranked_candidates() -> None:
    payload = {
        "candidate": "Jane",
        "criterion_scores": [{"criterion": "A", "score": 80}],
        "summary": "ok",
    }
    llm = StubLLM([json.dumps(payload)])
    result = screen_cvs(llm, "p", [_cv("jane.pdf")], criteria=[JobCriterion(name="A", weight=1.0)])  # type: ignore[arg-type]
    markdown = render_screening_markdown(result)
    assert "Jane" in markdown
    assert "Rank" in markdown


def test_render_screening_excel_writes_readable_file(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    payload = {
        "candidate": "Jane Doe",
        "criterion_scores": [{"criterion": "A", "score": 80}],
        "summary": "Strong",
        "gaps": ["g"],
    }
    llm = StubLLM([json.dumps(payload)])
    result = screen_cvs(llm, "p", [_cv("jane.docx")], criteria=[JobCriterion(name="A", weight=1.0)])  # type: ignore[arg-type]
    out = render_screening_excel(result, tmp_path, AppConfig())

    assert out.is_file()
    worksheet = load_workbook(out).active
    assert str(worksheet["A1"].value).startswith("CV Screening")
    found = any(cell.value == "Jane Doe" for row in worksheet.iter_rows() for cell in row)
    assert found
