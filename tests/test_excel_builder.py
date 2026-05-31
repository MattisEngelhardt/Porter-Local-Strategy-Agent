"""Tests for the Excel builder (core/excel_builder.py): formula integrity (N-10) + structure.

Workbooks are re-opened with openpyxl; the key assertions check that derived cells are genuine
Excel formulas (SUMPRODUCT/RANK/…), never hardcoded values — so they recalculate in MS Excel.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from core.config import AppConfig
from core.excel_builder import (
    ExcelBuildError,
    build_benchmark_table,
    build_business_case,
    build_decision_matrix,
    build_tracker,
)
from models.task import Language
from models.workbook import (
    BenchmarkData,
    BenchmarkRow,
    BenchmarkSource,
    BusinessCaseData,
    CaseAssumption,
    DecisionMatrixData,
    EntityScores,
    ScoringCriterion,
    TrackerData,
    TrackerItem,
)


def _data(language: Language = Language.EN) -> DecisionMatrixData:
    return DecisionMatrixData(
        title="M&A Target Screening",
        language=language,
        criteria=[
            ScoringCriterion(name="Technology Fit", weight=30, definition="Complementary tech"),
            ScoringCriterion(name="Market Access", weight=25, definition="New geography"),
            ScoringCriterion(name="Integration", weight=20, definition="Low complexity = high"),
            ScoringCriterion(name="Team Quality", weight=25, definition="Founder track record"),
        ],
        entities=[
            EntityScores(name="Dexory", scores=[4, 5, 3, 4], notes=["patents", "UK", "", "strong"]),
            EntityScores(name="Wandercraft", scores=[5, 3, 2, 5]),
            EntityScores(name="Exotec", scores=[3, 4, 4, 3]),
        ],
    )


def test_decision_matrix_has_three_tabs(tmp_path: Path) -> None:
    """E-1 has Summary_Matrix, Criteria_Guide, and Research_Notes tabs (SPEC §12)."""
    path = build_decision_matrix(_data(), AppConfig(), tmp_path)
    assert path.exists() and path.suffix == ".xlsx"
    wb = load_workbook(str(path))
    assert wb.sheetnames == ["Summary_Matrix", "Criteria_Guide", "Research_Notes"]


def test_decision_matrix_uses_real_formulas(tmp_path: Path) -> None:
    """Weighted score = SUMPRODUCT, rank = RANK, weights auto-sum — no hardcoded values (N-10)."""
    path = build_decision_matrix(_data(), AppConfig(), tmp_path)
    ws = load_workbook(str(path))["Summary_Matrix"]
    # 4 criteria → weighted score in column F (6), rank in column G (7); first entity at row 7.
    assert ws["F7"].value == "=SUMPRODUCT(B7:E7,$B$6:$E$6)"
    assert ws["G7"].value == "=RANK(F7,$F$7:$F$9,0)"
    assert ws["F6"].value == "=SUM(B6:E6)"
    # The last entity also has formulas (every row, not just the first).
    assert ws["F9"].value == "=SUMPRODUCT(B9:E9,$B$6:$E$6)"


def test_decision_matrix_no_cached_values(tmp_path: Path) -> None:
    """A data_only re-open returns None for formula cells (proves no hardcoded intermediates)."""
    path = build_decision_matrix(_data(), AppConfig(), tmp_path)
    ws = load_workbook(str(path), data_only=True)["Summary_Matrix"]
    assert ws["F7"].value is None  # never opened in Excel → a formula, not a literal


def test_decision_matrix_yellow_inputs_and_conditional_formatting(tmp_path: Path) -> None:
    """Weights + scores are yellow input cells; the score column has conditional formatting."""
    config = AppConfig()
    path = build_decision_matrix(_data(), config, tmp_path)
    ws = load_workbook(str(path))["Summary_Matrix"]
    yellow = config.output.colors.excel_input_cell.lstrip("#").upper()
    assert ws["B6"].fill.fgColor.rgb.endswith(yellow)  # a weight is a yellow input
    assert ws["B7"].fill.fgColor.rgb.endswith(yellow)  # a score is a yellow input
    assert ws["B6"].number_format == "0%"
    assert len(ws.conditional_formatting._cf_rules) >= 2  # colour scale + top-rank highlight


def test_decision_matrix_freeze_and_widths(tmp_path: Path) -> None:
    """Header + entity-name pane is frozen and column widths are set (not Excel default)."""
    path = build_decision_matrix(_data(), AppConfig(), tmp_path)
    ws = load_workbook(str(path))["Summary_Matrix"]
    assert ws.freeze_panes == "B7"
    assert ws.column_dimensions["A"].width and ws.column_dimensions["A"].width > 8.43


def test_decision_matrix_is_bilingual(tmp_path: Path) -> None:
    """German data renders German headers; English data renders English headers."""
    de = load_workbook(str(build_decision_matrix(_data(Language.DE), AppConfig(), tmp_path)))
    en = load_workbook(str(build_decision_matrix(_data(Language.EN), AppConfig(), tmp_path)))
    assert de["Summary_Matrix"]["F5"].value == "Gewichteter Score"
    assert en["Summary_Matrix"]["F5"].value == "Weighted Score"


def test_decision_matrix_requires_criteria(tmp_path: Path) -> None:
    """Building a matrix with no criteria fails fast (ExcelBuildError)."""
    empty = DecisionMatrixData(title="x", language=Language.EN, criteria=[], entities=[])
    with pytest.raises(ExcelBuildError):
        build_decision_matrix(empty, AppConfig(), tmp_path)


# ----------------------------------------------------------------- E-2 Benchmark Table
def _benchmark() -> BenchmarkData:
    return BenchmarkData(
        title="Humanoid Funding Benchmark",
        language=Language.EN,
        metrics=["Founded", "HQ", "Total Funding", "Lead Investor"],
        rows=[
            BenchmarkRow(name="Figure", values=["2022", "Sunnyvale", "$854M", "Microsoft"]),
            BenchmarkRow(name="1X", values=["2014", "Norway", "$125M", "EQT"]),
        ],
        sources=[
            BenchmarkSource(
                entity="Figure",
                metric="Total Funding",
                value="$854M",
                url="https://techcrunch.com",
                date="2024",
                confidence="High",
            )
        ],
    )


def test_benchmark_table_has_excel_table_and_sources(tmp_path: Path) -> None:
    """E-2 wraps the data in an Excel Table (auto-filter) and adds a Sources tab (SPEC §12)."""
    path = build_benchmark_table(_benchmark(), AppConfig(), tmp_path)
    wb = load_workbook(str(path))
    assert wb.sheetnames == ["Benchmark_Table", "Sources"]
    ws = wb["Benchmark_Table"]
    assert "BenchmarkTable" in ws.tables  # an Excel Table → sortable/filterable
    assert ws["A5"].value == "Company" and ws["E5"].value == "Lead Investor"
    assert ws["A6"].value == "Figure"
    assert wb["Sources"]["D6"].value == "https://techcrunch.com"


def test_benchmark_requires_metrics(tmp_path: Path) -> None:
    """A benchmark with no metric columns fails fast."""
    empty = BenchmarkData(title="x", language=Language.EN, metrics=[], rows=[])
    with pytest.raises(ExcelBuildError):
        build_benchmark_table(empty, AppConfig(), tmp_path)


# ----------------------------------------------------------------- E-4 Tracker Dashboard
def _tracker() -> TrackerData:
    return TrackerData(
        title="M&A Pipeline",
        language=Language.EN,
        items=[
            TrackerItem(name="Dexory", category="Warehouse", status="Active", priority="High"),
            TrackerItem(name="Exotec", category="Logistics", status="On Hold", priority="Medium"),
        ],
    )


def test_tracker_three_tabs_dropdowns_and_dashboard_formula(tmp_path: Path) -> None:
    """E-4 has Dashboard/Tracker/Archive; dropdowns + a COUNTIF dashboard formula (SPEC §12)."""
    path = build_tracker(_tracker(), AppConfig(), tmp_path)
    wb = load_workbook(str(path))
    assert wb.sheetnames == ["Dashboard", "Tracker", "Archive"]
    tracker = wb["Tracker"]
    # Status + Priority data-validation dropdowns.
    assert len(tracker.data_validations.dataValidation) == 2
    formulas = {dv.formula1 for dv in tracker.data_validations.dataValidation}
    assert any("Active" in f for f in formulas) and any("High" in f for f in formulas)
    # Conditional formatting present (status colours + priority).
    assert len(tracker.conditional_formatting._cf_rules) >= 2
    # Dashboard stat is a real formula linked to the Tracker tab (not a hardcoded count).
    dashboard = wb["Dashboard"]
    active = dashboard["B4"].value
    assert isinstance(active, str) and active.startswith("=COUNTIF(Tracker!")


def test_tracker_prepopulates_items(tmp_path: Path) -> None:
    """Provided items are written into the Tracker tab (pre-populated from research)."""
    path = build_tracker(_tracker(), AppConfig(), tmp_path)
    tracker = load_workbook(str(path))["Tracker"]
    assert tracker["A4"].value == "Dexory"
    assert tracker["C4"].value == "Active"


# ----------------------------------------------------------------- E-3 Business Case Model
def _business_case(language: Language = Language.EN) -> BusinessCaseData:
    return BusinessCaseData(
        title="Japan Expansion",
        language=language,
        years=3,
        investment=2_000_000,
        revenue_year1=1_500_000,
        revenue_growth=0.4,
        opex_year1=900_000,
        opex_growth=0.15,
        discount_rate=0.12,
        assumptions=[
            CaseAssumption(
                name="Market size",
                value=500_000_000,
                unit="EUR",
                source="JETRO 2026",
                confidence="Medium",
            )
        ],
        bottom_line="Positive NPV by year 3; recommend phased entry.",
    )


def test_business_case_has_five_tabs(tmp_path: Path) -> None:
    """E-3 has the 5 SPEC §12 tabs (Summary/Assumptions/Projections/Scenarios/Sources)."""
    path = build_business_case(_business_case(), AppConfig(), tmp_path)
    assert path.suffix == ".xlsx"
    wb = load_workbook(str(path))
    assert wb.sheetnames == ["Summary", "Assumptions", "Projections", "Scenarios", "Sources"]


def test_business_case_assumptions_are_yellow_inputs(tmp_path: Path) -> None:
    """All core drivers live on the Assumptions tab as yellow input cells."""
    config = AppConfig()
    path = build_business_case(_business_case(), config, tmp_path)
    a = load_workbook(str(path))["Assumptions"]
    yellow = config.output.colors.excel_input_cell.lstrip("#").upper()
    assert a["B4"].value == 2_000_000  # investment
    assert a["B5"].value == 1_500_000  # revenue year 1
    assert a["B4"].fill.fgColor.rgb.endswith(yellow)
    assert a["B5"].fill.fgColor.rgb.endswith(yellow)


def test_business_case_projections_reference_assumptions(tmp_path: Path) -> None:
    """Projections are formulas referencing the Assumptions tab — no hardcoded numbers (N-10)."""
    path = build_business_case(_business_case(), AppConfig(), tmp_path)
    p = load_workbook(str(path))["Projections"]
    assert p["B7"].value == "=Assumptions!$B$5"  # year 1 revenue = assumption
    assert p["B8"].value == "=B7*(1+Assumptions!$B$6)"  # year 2 grows by the growth assumption
    assert p["D7"].value == "=B7-C7"  # EBITDA = revenue - opex
    assert p["F7"].value == "=F6+E7"  # cumulative cash flow


def test_business_case_summary_npv_irr_formulas(tmp_path: Path) -> None:
    """The Summary tab computes NPV and IRR via Excel functions over the projected cash flows."""
    path = build_business_case(_business_case(), AppConfig(), tmp_path)
    s = load_workbook(str(path))["Summary"]
    npv = s["B6"].value
    irr = s["B7"].value
    assert isinstance(npv, str) and npv.startswith("=-Assumptions!$B$4+NPV(")
    assert isinstance(irr, str) and "IRR(Projections!" in irr


def test_business_case_scenarios_use_yellow_multipliers(tmp_path: Path) -> None:
    """Each scenario's NPV is a formula scaling projections by yellow multiplier inputs."""
    config = AppConfig()
    path = build_business_case(_business_case(), config, tmp_path)
    sc = load_workbook(str(path))["Scenarios"]
    yellow = config.output.colors.excel_input_cell.lstrip("#").upper()
    assert sc["B7"].fill.fgColor.rgb.endswith(yellow)  # revenue factor is a yellow input
    assert isinstance(sc["D7"].value, str) and sc["D7"].value.startswith("=-Assumptions!$B$4+NPV(")


def test_business_case_no_cached_values(tmp_path: Path) -> None:
    """A data_only re-open returns None for the headline metrics (proves pure formulas, N-10)."""
    path = build_business_case(_business_case(), AppConfig(), tmp_path)
    s = load_workbook(str(path), data_only=True)["Summary"]
    assert s["B6"].value is None  # NPV is a formula, never a hardcoded value
